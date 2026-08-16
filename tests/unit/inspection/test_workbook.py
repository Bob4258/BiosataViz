from pathlib import Path

import openpyxl
import pytest

from biostatviz.io import ExcelSheetNotFoundError, UnsupportedTableFormatError


def _make_batch_workbook(path: Path):
    wb = openpyxl.Workbook()
    clean = wb.active; clean.title = "Clean"; clean.append(["sample", "value"]); clean.append(["S1", 1]); clean.append(["S2", 2])
    merged = wb.create_sheet("Merged"); merged["A1"] = "time"; merged.merge_cells("B1:C1"); merged["B1"] = "Group"; merged.append(["t0", 1, 2]); merged.append(["t1", 3, 4])
    multi = wb.create_sheet("Multi"); multi.append(["a", "b", None, "x", "y"]); multi.append([1, 2, None, 3, 4]); multi.append([5, 6, None, 7, 8])
    last = wb.create_sheet("Last"); last.append(["sample", "value"]); last.append(["L1", 5]); last.append(["L2", 6])
    wb.save(path)


def test_inspect_workbook_defaults_to_all_sheets_in_workbook_order(tmp_path: Path):
    from biostatviz.inspection import inspect_workbook
    path = tmp_path / "batch.xlsx"; _make_batch_workbook(path)
    report = inspect_workbook(path)
    assert report.sheet_names == ("Clean", "Merged", "Multi", "Last")
    assert len(report.sheets) == 4
    assert report.sheets[0].action_required_count == 0
    assert report.sheets[1].action_required_count >= 1
    assert report.sheets[2].action_required_count >= 1
    assert report.sheets[3].report.inspected_sheet == "Last"
    assert report.requires_user_input is True


def test_inspect_workbook_explicit_subset_preserves_caller_order(tmp_path: Path):
    from biostatviz.inspection import inspect_workbook
    path = tmp_path / "batch.xlsx"; _make_batch_workbook(path)
    assert inspect_workbook(path, sheets=["Last", "Clean"]).sheet_names == ("Last", "Clean")


def test_inspect_workbook_resolves_integer_indices_and_rejects_duplicates(tmp_path: Path):
    from biostatviz.inspection import inspect_workbook
    path = tmp_path / "batch.xlsx"; _make_batch_workbook(path)
    assert inspect_workbook(path, sheets=[3, 0]).sheet_names == ("Last", "Clean")
    with pytest.raises(ValueError, match="Duplicate sheet selector"):
        inspect_workbook(path, sheets=["Clean", 0])


def test_inspect_workbook_rejects_missing_sheet_and_csv(tmp_path: Path):
    from biostatviz.inspection import inspect_workbook
    path = tmp_path / "batch.xlsx"; _make_batch_workbook(path)
    with pytest.raises(ExcelSheetNotFoundError): inspect_workbook(path, sheets=["Missing"])
    csv_path = tmp_path / "data.csv"; csv_path.write_text("a,b\n1,2\n", encoding="utf-8")
    with pytest.raises(UnsupportedTableFormatError): inspect_workbook(csv_path)


def test_workbook_derived_counts_and_source_immutability(tmp_path: Path):
    from biostatviz.inspection import inspect_workbook
    path = tmp_path / "batch.xlsx"; _make_batch_workbook(path); before = path.read_bytes(); report = inspect_workbook(path)
    merged = next(item for item in report.sheets if item.sheet_name == "Merged"); multi = next(item for item in report.sheets if item.sheet_name == "Multi")
    assert merged.region_count == 1
    assert merged.merged_header_count == 1
    assert multi.region_count == 2
    assert path.read_bytes() == before
