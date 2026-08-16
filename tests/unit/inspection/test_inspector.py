from pathlib import Path

import openpyxl

from biostatviz.inspection import InspectionSeverity


def test_clean_csv_has_no_blocking_issues(tmp_path: Path):
    from biostatviz.inspection import inspect_table

    path = tmp_path / "clean.csv"
    path.write_text("sample,value\nS1,1\nS2,2\n", encoding="utf-8")
    report = inspect_table(path)
    assert report.source_format == "csv"
    assert report.shape_preview == (3, 2)
    assert report.requires_user_input is False


def test_xlsx_multiple_regions_require_explicit_selection(tmp_path: Path):
    from biostatviz.inspection import inspect_table

    path = tmp_path / "regions.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    for row in (("a", "b", None, None, "x", "y"), (1, 2, None, None, 3, 4), (5, 6, None, None, 7, 8)):
        ws.append(row)
    wb.save(path)

    report = inspect_table(path, sheet_name="Data")
    issue = next(i for i in report.issues if i.code == "MULTIPLE_TABLE_REGIONS")
    assert issue.severity is InspectionSeverity.ACTION_REQUIRED
    assert issue.candidates == ("region:1", "region:2")
    assert [r.a1_range for r in report.regions] == ["A1:B3", "E1:F3"]
    assert report.selected_region is None


def test_selected_region_limits_inspection_and_detects_merged_header(tmp_path: Path):
    from biostatviz.inspection import inspect_table

    path = tmp_path / "merged.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.merge_cells("A1:C1")
    ws["A1"] = "Group"
    ws.append([None, None, None])
    ws["A2"], ws["B2"], ws["C2"] = "r1", "r2", "r3"
    ws["A3"], ws["B3"], ws["C3"] = 1, 2, 3
    ws["A4"], ws["B4"], ws["C4"] = 4, 5, 6
    wb.save(path)

    report = inspect_table(path, sheet_name="Data", region="A1:C4")
    issue = next(i for i in report.issues if i.code == "MERGED_HEADER_DETECTED")
    assert issue.severity is InspectionSeverity.ACTION_REQUIRED
    assert issue.observed == ("A1:C1=Group",)
    assert report.selected_region is not None
    assert report.selected_region.a1_range == "A1:C4"
    assert not any(i.code in {"BLANK_COLUMN", "MIXED_NUMERIC_TEXT"} for i in report.issues)


def test_inspection_does_not_modify_source_bytes(tmp_path: Path):
    from biostatviz.inspection import inspect_table

    path = tmp_path / "data.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["a", "b"])
    ws.append([1, 2])
    wb.save(path)
    before = path.read_bytes()
    inspect_table(path, sheet_name=0)
    assert path.read_bytes() == before


def test_style_only_extent_does_not_trigger_false_region_scan_truncation(tmp_path: Path):
    from openpyxl.styles import PatternFill
    from biostatviz.inspection import inspect_table

    path = tmp_path / "style-residue.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"], ws["B1"] = "sample", "value"
    ws["A2"], ws["B2"] = "S1", 1
    ws["A3"], ws["B3"] = "S2", 2
    ws.merge_cells("A3000:B3000")  # blank merge residue
    ws["J5000"].fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    wb.save(path)

    report = inspect_table(path, sheet_name="Data")
    assert not any(i.code == "REGION_SCAN_TRUNCATED" for i in report.issues)
