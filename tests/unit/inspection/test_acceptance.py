from pathlib import Path

import openpyxl

from biostatviz.inspection import inspect_table, inspect_workbook
from biostatviz.io import load_table


def test_inspect_then_explicit_load_preserves_scientific_tokens(tmp_path: Path):
    path = tmp_path / "experiment.csv"
    path.write_text("Experiment note,,\nsample,group,intensity\nS1,A,101\nS2,A,NA\nS3,B,145\n", encoding="utf-8")
    report = inspect_table(path)
    assert report.requires_user_input is True
    option = next(item for item in report.candidate_loading_options if item.header == 1)
    loaded = load_table(path, options=option)
    assert loaded.data.columns.tolist() == ["sample", "group", "intensity"]
    assert loaded.data["intensity"].tolist() == ["101", "NA", "145"]


def test_workbook_batch_inspection_combines_independent_sheet_results(tmp_path: Path):
    from openpyxl.styles import PatternFill
    path = tmp_path / "batch.xlsx"
    wb = openpyxl.Workbook()
    first = wb.active; first.title = "Set A"; first.append(["sample", "value"]); first.append(["A1", 1]); first.append(["A2", 2])
    second = wb.create_sheet("Set B"); second["A1"] = "time"; second.merge_cells("B1:C1"); second["B1"] = "Group"; second.append(["t0", 1, 2]); second.append(["t1", 3, 4]); second.merge_cells("A3000:B3000"); second["J5000"].fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    third = wb.create_sheet("Set C"); third.append(["x", "y", None, "u", "v"]); third.append([1, 2, None, 3, 4]); third.append([5, 6, None, 7, 8]); wb.save(path)
    before = path.read_bytes(); report = inspect_workbook(path)
    assert report.sheet_names == ("Set A", "Set B", "Set C")
    assert report.sheets[0].action_required_count == 0
    assert report.sheets[1].merged_header_count == 1
    assert report.sheets[2].region_count == 2
    assert path.read_bytes() == before


def test_clean_example_csv_remains_nonblocking():
    root = Path(__file__).resolve().parents[3]
    report = inspect_table(root / "examples" / "data" / "two_group_gfp.csv")
    assert report.requires_user_input is False
