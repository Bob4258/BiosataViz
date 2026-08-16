from pathlib import Path

import openpyxl
import pytest

from biostatviz.inspection import InspectionReadError
from biostatviz.inspection.preview import read_xlsx_preview


def test_xlsx_preview_preserves_sheets_raw_tokens_and_merged_ranges(tmp_path: Path):
    path = tmp_path / "book.xlsx"
    wb = openpyxl.Workbook()
    raw = wb.active
    raw.title = "Raw"
    raw.append(["sample", "value", None])
    raw.append(["S1", "NA", None])
    raw.merge_cells("B4:C4")
    raw["B4"] = "Group"
    notes = wb.create_sheet("Notes")
    notes.append(["free text"])
    wb.save(path)
    preview = read_xlsx_preview(path, max_rows=20, max_columns=20)
    assert tuple(s.summary.name for s in preview.sheets) == ("Raw", "Notes")
    assert preview.sheets[0].table.rows[1][1] == "NA"
    assert preview.sheets[0].table.merged_ranges[0].a1_range == "B4:C4"
    assert preview.sheets[0].table.merged_ranges[0].anchor_value == "Group"


def test_xlsx_preview_marks_only_table_like_sheet_plausible(tmp_path: Path):
    path = tmp_path / "book.xlsx"
    wb = openpyxl.Workbook()
    data = wb.active
    data.title = "Data"
    data.append(["sample", "value"])
    data.append(["S1", 1])
    notes = wb.create_sheet("Notes")
    notes["A1"] = "one cell"
    wb.save(path)
    preview = read_xlsx_preview(path)
    assert [s.summary.plausible_table for s in preview.sheets] == [True, False]


def test_xlsx_preview_wraps_corrupt_file(tmp_path: Path):
    path = tmp_path / "broken.xlsx"
    path.write_bytes(b"not an xlsx")
    with pytest.raises(InspectionReadError, match="broken.xlsx"):
        read_xlsx_preview(path)


def test_xlsx_preview_ignores_empty_merged_ranges_and_style_only_extent(tmp_path: Path):
    from openpyxl.styles import PatternFill
    path = tmp_path / "residue.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["B2"] = "title"
    ws.merge_cells("C3:G3"); ws["C3"] = "Group A"
    ws.merge_cells("H3:L3"); ws["H3"] = "Group B"
    for row in range(4, 8):
        for col in range(2, 13):
            ws.cell(row, col).value = row * col
    ws.merge_cells("M26:Q26")
    ws.merge_cells("F76:H76")
    ws["AB108"].fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    wb.save(path)
    preview = read_xlsx_preview(path, max_rows=2000, max_columns=256)
    sheet = preview.sheets[0]
    assert tuple(r.a1_range for r in sheet.table.merged_ranges) == ("C3:G3", "H3:L3")
    assert sheet.table.apparent_rows == 7
    assert sheet.table.apparent_columns == 12


def test_xlsx_preview_treats_zero_false_and_formula_anchors_as_meaningful(tmp_path: Path):
    path = tmp_path / "anchors.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.merge_cells("A1:B1"); ws["A1"] = 0
    ws.merge_cells("A2:B2"); ws["A2"] = False
    ws.merge_cells("A3:B3"); ws["A3"] = "=1+1"
    wb.save(path)
    preview = read_xlsx_preview(path)
    assert [r.anchor_value for r in preview.sheets[0].table.merged_ranges] == [0, False, "=1+1"]
