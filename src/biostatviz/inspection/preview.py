"""Raw bounded previews used before pandas can coerce scientific tokens."""

from __future__ import annotations

import csv
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable


from .errors import InspectionReadError
from .models import SheetSummary

DELIMITER_CANDIDATES = (",", "\t", ";", "|")


@dataclass(frozen=True, slots=True)
class MergedRange:
    min_row: int
    max_row: int
    min_col: int
    max_col: int
    anchor_value: object

    @property
    def a1_range(self) -> str:
        from openpyxl.utils import get_column_letter

        return (
            f"{get_column_letter(self.min_col)}{self.min_row}:"
            f"{get_column_letter(self.max_col)}{self.max_row}"
        )


@dataclass(frozen=True, slots=True)
class TablePreview:
    rows: tuple[tuple[object, ...], ...]
    delimiter: str | None = None
    delimiter_candidates: tuple[str, ...] = ()
    sheet_name: str | None = None
    merged_ranges: tuple[MergedRange, ...] = ()
    origin_row: int = 1
    origin_col: int = 1
    apparent_rows: int = 0
    apparent_columns: int = 0


@dataclass(frozen=True, slots=True)
class SheetPreview:
    summary: SheetSummary
    table: TablePreview


@dataclass(frozen=True, slots=True)
class WorkbookPreview:
    sheets: tuple[SheetPreview, ...]


def _blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _parse_lines(lines: list[str], delimiter: str) -> list[tuple[str, ...]]:
    return [tuple(row) for row in csv.reader(lines, delimiter=delimiter)]


def _candidate_score(lines: list[str], delimiter: str) -> tuple[float, int] | None:
    rows = _parse_lines(lines, delimiter)
    nonblank = [row for row in rows if any(cell.strip() for cell in row)]
    if not nonblank:
        return None
    widths = [len(row) for row in nonblank]
    modal_width, modal_count = Counter(widths).most_common(1)[0]
    ratio = modal_count / len(nonblank)
    if modal_width < 2 or ratio < 0.80:
        return None
    return (ratio, modal_width)


def read_csv_preview(
    path: Path,
    *,
    encoding: str = "utf-8-sig",
    delimiter: str | None = None,
    max_rows: int = 50,
) -> TablePreview:
    """Read raw CSV text without pandas NA/type coercion."""

    try:
        with Path(path).open("r", encoding=encoding, newline="") as handle:
            lines: list[str] = []
            for _ in range(max_rows):
                line = handle.readline()
                if line == "":
                    break
                lines.append(line)
    except (UnicodeDecodeError, OSError) as exc:
        raise InspectionReadError(f"Failed to preview CSV file: {path}") from exc

    try:
        if delimiter is not None:
            rows = _parse_lines(lines, delimiter)
            return TablePreview(
                rows=tuple(rows),
                delimiter=delimiter,
                delimiter_candidates=(delimiter,),
                apparent_rows=len(rows),
                apparent_columns=max((len(r) for r in rows), default=0),
            )

        scored: list[tuple[str, tuple[float, int]]] = []
        for candidate in DELIMITER_CANDIDATES:
            score = _candidate_score(lines, candidate)
            if score is not None:
                scored.append((candidate, score))

        if not scored:
            chosen = ","
            rows = _parse_lines(lines, chosen)
            return TablePreview(
                rows=tuple(rows),
                delimiter=chosen,
                delimiter_candidates=(),
                apparent_rows=len(rows),
                apparent_columns=max((len(r) for r in rows), default=0),
            )

        best_score = max(score for _, score in scored)
        tied = tuple(candidate for candidate, score in scored if score == best_score)
        if len(tied) == 1:
            chosen = tied[0]
            rows = _parse_lines(lines, chosen)
            return TablePreview(
                rows=tuple(rows),
                delimiter=chosen,
                delimiter_candidates=tied,
                apparent_rows=len(rows),
                apparent_columns=max((len(r) for r in rows), default=0),
            )

        rows = _parse_lines(lines, tied[0])
        return TablePreview(
            rows=tuple(rows),
            delimiter=None,
            delimiter_candidates=tied,
            apparent_rows=len(rows),
            apparent_columns=max((len(r) for r in rows), default=0),
        )
    except csv.Error as exc:
        raise InspectionReadError(f"Failed to preview CSV file: {path}") from exc


def _meaningful_merged_ranges(ws) -> tuple[MergedRange, ...]:
    ranges: list[MergedRange] = []
    for cell_range in ws.merged_cells.ranges:
        anchor = ws.cell(cell_range.min_row, cell_range.min_col).value
        if _blank(anchor):
            continue
        ranges.append(
            MergedRange(
                min_row=cell_range.min_row,
                max_row=cell_range.max_row,
                min_col=cell_range.min_col,
                max_col=cell_range.max_col,
                anchor_value=anchor,
            )
        )
    ranges.sort(key=lambda item: (item.min_row, item.min_col, item.max_row, item.max_col))
    return tuple(ranges)


def _meaningful_bounds(ws, merged_ranges: tuple[MergedRange, ...]) -> tuple[int, int]:
    max_row = 0
    max_col = 0
    for cell in ws._cells.values():
        if _blank(cell.value):
            continue
        max_row = max(max_row, cell.row)
        max_col = max(max_col, cell.column)
    for merged in merged_ranges:
        max_row = max(max_row, merged.max_row)
        max_col = max(max_col, merged.max_col)
    return max_row, max_col


def _build_sheet_preview(ws, *, max_rows: int, max_columns: int) -> SheetPreview:
    merged_ranges = _meaningful_merged_ranges(ws)
    effective_rows, effective_columns = _meaningful_bounds(ws, merged_ranges)
    row_limit = min(effective_rows, max_rows)
    column_limit = min(effective_columns, max_columns)
    rows = tuple(
        tuple(ws.cell(row=row, column=column).value for column in range(1, column_limit + 1))
        for row in range(1, row_limit + 1)
    )
    nonempty_rows = sum(any(not _blank(value) for value in row) for row in rows)
    nonempty_columns = {
        column_index
        for row in rows
        for column_index, value in enumerate(row, start=1)
        if not _blank(value)
    }
    return SheetPreview(
        summary=SheetSummary(
            name=ws.title,
            non_empty_rows=nonempty_rows,
            max_columns=max(nonempty_columns, default=0),
            plausible_table=nonempty_rows >= 2 and len(nonempty_columns) >= 2,
        ),
        table=TablePreview(
            rows=rows,
            sheet_name=ws.title,
            merged_ranges=merged_ranges,
            apparent_rows=effective_rows,
            apparent_columns=effective_columns,
        ),
    )


def read_xlsx_preview(
    path: Path,
    *,
    max_rows: int = 50,
    max_columns: int = 50,
) -> WorkbookPreview:
    """Read bounded XLSX raw values and merged ranges without pandas."""
    try:
        import openpyxl

        workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    except Exception as exc:
        raise InspectionReadError(f"Failed to preview XLSX file: {path}") from exc
    try:
        return WorkbookPreview(
            sheets=tuple(
                _build_sheet_preview(ws, max_rows=max_rows, max_columns=max_columns)
                for ws in workbook.worksheets
            )
        )
    finally:
        workbook.close()
