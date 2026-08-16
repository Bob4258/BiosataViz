"""Immutable data contracts for BioStatViz inspection."""

from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from typing import Literal

from biostatviz.io import LoadingOptions


class InspectionSeverity(str, Enum):
    INFO = "INFO"
    WARNING = "WARNING"
    ACTION_REQUIRED = "ACTION_REQUIRED"


@dataclass(frozen=True, slots=True)
class InspectionIssue:
    code: str
    severity: InspectionSeverity
    message: str
    location: str | None = None
    observed: tuple[str, ...] = ()
    candidates: tuple[str, ...] = ()
    suggested_action: str | None = None


@dataclass(frozen=True, slots=True)
class SheetSummary:
    name: str
    non_empty_rows: int
    max_columns: int
    plausible_table: bool


@dataclass(frozen=True, slots=True)
class TableRegion:
    region_id: str
    sheet_name: str
    min_row: int
    max_row: int
    min_col: int
    max_col: int
    non_empty_cells: int
    preview_values: tuple[str, ...] = ()
    merged_ranges: tuple[str, ...] = ()

    @property
    def a1_range(self) -> str:
        from openpyxl.utils import get_column_letter

        return (
            f"{get_column_letter(self.min_col)}{self.min_row}:"
            f"{get_column_letter(self.max_col)}{self.max_row}"
        )


@dataclass(frozen=True, slots=True)
class InspectionReport:
    source_path: Path
    source_format: Literal["csv", "xlsx"]
    sheets: tuple[SheetSummary, ...] = ()
    inspected_sheet: str | int | None = None
    shape_preview: tuple[int, int] = (0, 0)
    issues: tuple[InspectionIssue, ...] = ()
    candidate_loading_options: tuple[LoadingOptions, ...] = ()
    regions: tuple[TableRegion, ...] = ()
    selected_region: TableRegion | None = None

    @property
    def requires_user_input(self) -> bool:
        return any(issue.severity is InspectionSeverity.ACTION_REQUIRED for issue in self.issues)

    @property
    def has_warnings(self) -> bool:
        return any(issue.severity is InspectionSeverity.WARNING for issue in self.issues)


@dataclass(frozen=True, slots=True)
class WorkbookSheetResult:
    """One worksheet's independent inspection result inside a workbook batch."""

    sheet_name: str
    report: InspectionReport

    @property
    def info_count(self) -> int:
        return sum(issue.severity is InspectionSeverity.INFO for issue in self.report.issues)

    @property
    def warning_count(self) -> int:
        return sum(issue.severity is InspectionSeverity.WARNING for issue in self.report.issues)

    @property
    def action_required_count(self) -> int:
        return sum(
            issue.severity is InspectionSeverity.ACTION_REQUIRED for issue in self.report.issues
        )

    @property
    def region_count(self) -> int:
        return len(self.report.regions)

    @property
    def merged_header_count(self) -> int:
        ranges: set[str] = set()
        for issue in self.report.issues:
            if issue.code != "MERGED_HEADER_DETECTED":
                continue
            for value in issue.observed:
                a1_range, _, _ = value.partition("=")
                if a1_range:
                    ranges.add(a1_range)
        return len(ranges)


@dataclass(frozen=True, slots=True)
class WorkbookInspectionReport:
    """Deterministic summary of independently inspected worksheets."""

    source_path: Path
    sheets: tuple[WorkbookSheetResult, ...]

    @property
    def sheet_names(self) -> tuple[str, ...]:
        return tuple(item.sheet_name for item in self.sheets)

    @property
    def has_warnings(self) -> bool:
        return any(item.report.has_warnings for item in self.sheets)

    @property
    def requires_user_input(self) -> bool:
        return any(item.report.requires_user_input for item in self.sheets)
