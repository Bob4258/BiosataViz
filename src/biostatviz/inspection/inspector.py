"""Top-level deterministic table inspection orchestration."""

from __future__ import annotations

from dataclasses import replace
from pathlib import Path
from openpyxl.utils.cell import range_boundaries
from biostatviz.io import ExcelSheetNotFoundError, LoadingOptions, TableNotFoundError, UnsupportedTableFormatError
from .checks import inspect_quality, inspect_structure
from .errors import RegionSelectionError
from .models import InspectionIssue, InspectionReport, InspectionSeverity, TableRegion
from .preview import TablePreview, read_csv_preview, read_xlsx_preview
from .regions import discover_regions

SUPPORTED_SUFFIXES={".csv",".xlsx"}


def _validate_path(path):
    source_path=Path(path).expanduser().resolve()
    if not source_path.exists(): raise TableNotFoundError(f"Table file does not exist: {source_path}")
    suffix=source_path.suffix.lower()
    if suffix not in SUPPORTED_SUFFIXES: raise UnsupportedTableFormatError(f"Unsupported table format '{suffix or '<none>'}'. Supported formats are .csv and .xlsx.")
    return source_path


def _dedupe_options(options):
    seen=set(); result=[]
    for option in options:
        if option not in seen: seen.add(option); result.append(option)
    return tuple(result)


def _select_sheet(workbook,sheet_name):
    sheets=workbook.sheets; summaries=tuple(s.summary for s in sheets)
    if sheet_name is not None:
        if isinstance(sheet_name,int):
            if sheet_name<0 or sheet_name>=len(sheets): raise ExcelSheetNotFoundError(f"Excel sheet not found: {sheet_name!r}")
            return sheets[sheet_name],summaries,sheet_name,()
        for sheet in sheets:
            if sheet.summary.name==sheet_name: return sheet,summaries,sheet_name,()
        raise ExcelSheetNotFoundError(f"Excel sheet not found: {sheet_name!r}")
    if len(sheets)==1: return sheets[0],summaries,0,()
    plausible=[s for s in sheets if s.summary.plausible_table]
    if len(plausible)==1: return plausible[0],summaries,plausible[0].summary.name,()
    if len(plausible)>=2:
        names=tuple(s.summary.name for s in plausible)
        return None,summaries,None,(InspectionIssue(code="MULTIPLE_PLAUSIBLE_SHEETS",severity=InspectionSeverity.ACTION_REQUIRED,message="Multiple worksheets contain plausible tables.",candidates=names,suggested_action="Choose a worksheet explicitly."),)
    names=tuple(s.summary.name for s in sheets)
    return None,summaries,None,(InspectionIssue(code="NO_PLAUSIBLE_TABLE_SHEET",severity=InspectionSeverity.ACTION_REQUIRED,message="No worksheet is clearly table-like from the bounded preview.",candidates=names,suggested_action="Choose a worksheet explicitly."),)


def _crop_preview(preview,region):
    rows=[]
    for source_row in range(region.min_row,region.max_row+1):
        ro=source_row-preview.origin_row; source=preview.rows[ro] if 0<=ro<len(preview.rows) else (); vals=[]
        for source_col in range(region.min_col,region.max_col+1):
            co=source_col-preview.origin_col; vals.append(source[co] if 0<=co<len(source) else None)
        rows.append(tuple(vals))
    overlapping=tuple(m for m in preview.merged_ranges if not (m.max_row<region.min_row or m.min_row>region.max_row or m.max_col<region.min_col or m.min_col>region.max_col))
    return TablePreview(rows=tuple(rows),sheet_name=preview.sheet_name,merged_ranges=overlapping,origin_row=region.min_row,origin_col=region.min_col,apparent_rows=region.max_row,apparent_columns=region.max_col)


def _count_nonblank(preview):
    return sum(value is not None and not (isinstance(value,str) and not value.strip()) for row in preview.rows for value in row)


def _resolve_region(selector,*,preview,regions,sheet_name):
    if isinstance(selector,TableRegion):
        if selector.sheet_name!=sheet_name: raise RegionSelectionError("Selected region does not belong to the selected worksheet.")
        return selector
    for region in regions:
        if selector==region.region_id or selector.upper()==region.a1_range.upper(): return region
    try: min_col,min_row,max_col,max_row=range_boundaries(selector)
    except Exception as exc: raise RegionSelectionError(f"Invalid region selector: {selector!r}") from exc
    if min_row<1 or min_col<1 or max_row<min_row or max_col<min_col: raise RegionSelectionError(f"Invalid region selector: {selector!r}")
    if max_row>preview.apparent_rows or max_col>preview.apparent_columns: raise RegionSelectionError(f"Region is outside the selected worksheet: {selector!r}")
    candidate=TableRegion(region_id="selected",sheet_name=sheet_name,min_row=min_row,max_row=max_row,min_col=min_col,max_col=max_col,non_empty_cells=0)
    return replace(candidate,non_empty_cells=_count_nonblank(_crop_preview(preview,candidate)))


def _merged_header_issue(preview,region):
    evidence=[]
    for merged in preview.merged_ranges:
        if merged.max_col-merged.min_col+1<2: continue
        if merged.min_row<region.min_row or merged.min_row>region.min_row+2: continue
        if merged.max_row<region.min_row or merged.min_row>region.max_row: continue
        if merged.max_col<region.min_col or merged.min_col>region.max_col: continue
        value=merged.anchor_value
        if value is None or (isinstance(value,str) and not value.strip()): continue
        evidence.append(f"{merged.a1_range}={value}")
    if not evidence: return None
    return InspectionIssue(code="MERGED_HEADER_DETECTED",severity=InspectionSeverity.ACTION_REQUIRED,message="The selected table uses merged cells in its leading header area.",location=region.a1_range,observed=tuple(evidence),suggested_action="Confirm how the merged structural labels should be handled.")


def _inspect_flat_preview(preview,*,source_path,source_format,summaries=(),inspected_sheet=None,base_options,regions=(),selected_region=None,prefix_issues=()):
    structure=inspect_structure(preview,base_options=base_options); issues=list(prefix_issues)+list(structure.issues)
    if structure.header_row is not None: issues.extend(inspect_quality(preview,header_row=structure.header_row))
    return InspectionReport(source_path=source_path,source_format=source_format,sheets=tuple(summaries),inspected_sheet=inspected_sheet,shape_preview=structure.shape_preview,issues=tuple(issues),candidate_loading_options=structure.candidate_loading_options,regions=tuple(regions),selected_region=selected_region)


def inspect_table(path,sheet_name=None,*,region=None,encoding="utf-8-sig",delimiter=None,max_rows=50,max_columns=50,max_region_rows=2000,max_region_columns=256):
    source_path=_validate_path(path); suffix=source_path.suffix.lower()
    if suffix==".csv":
        if region is not None: raise RegionSelectionError("Region selection is only valid for .xlsx files.")
        preview=read_csv_preview(source_path,encoding=encoding,delimiter=delimiter,max_rows=max_rows); base=LoadingOptions(delimiter=delimiter or ",",encoding=encoding)
        if preview.delimiter is None:
            issue=InspectionIssue(code="DELIMITER_AMBIGUOUS",severity=InspectionSeverity.ACTION_REQUIRED,message="Multiple delimiters are equally plausible.",candidates=preview.delimiter_candidates,suggested_action="Choose the delimiter explicitly.")
            return InspectionReport(source_path=source_path,source_format="csv",shape_preview=(len(preview.rows),max((len(r) for r in preview.rows),default=0)),issues=(issue,),candidate_loading_options=tuple(replace(base,delimiter=v) for v in preview.delimiter_candidates))
        return _inspect_flat_preview(preview,source_path=source_path,source_format="csv",base_options=replace(base,delimiter=preview.delimiter))
    workbook=read_xlsx_preview(source_path,max_rows=max_region_rows,max_columns=max_region_columns)
    selected,summaries,option_sheet,sheet_issues=_select_sheet(workbook,sheet_name)
    if selected is None:
        return InspectionReport(source_path=source_path,source_format="xlsx",sheets=summaries,issues=sheet_issues,candidate_loading_options=_dedupe_options([LoadingOptions(sheet_name=n) for n in sheet_issues[0].candidates]))
    label=selected.summary.name; preview=selected.table; base=LoadingOptions(sheet_name=option_sheet); regions,truncated=discover_regions(preview); prefix=[]
    if truncated: prefix.append(InspectionIssue(code="REGION_SCAN_TRUNCATED",severity=InspectionSeverity.WARNING,message="Table-region discovery was bounded before the apparent worksheet extent."))
    chosen=None
    if region is not None: chosen=_resolve_region(region,preview=preview,regions=regions,sheet_name=label)
    elif len(regions)>1:
        issue=InspectionIssue(code="MULTIPLE_TABLE_REGIONS",severity=InspectionSeverity.ACTION_REQUIRED,message="Multiple table-like regions were detected in the selected worksheet.",candidates=tuple(r.region_id for r in regions),suggested_action="Choose a table region explicitly.")
        return InspectionReport(source_path=source_path,source_format="xlsx",sheets=summaries,inspected_sheet=label,shape_preview=(len(preview.rows),max((len(r) for r in preview.rows),default=0)),issues=tuple(prefix+[issue]),regions=regions)
    elif len(regions)==1: chosen=regions[0]
    if chosen is not None:
        cropped=_crop_preview(preview,chosen); merged_issue=_merged_header_issue(cropped,chosen)
        if merged_issue is not None: return InspectionReport(source_path=source_path,source_format="xlsx",sheets=summaries,inspected_sheet=label,shape_preview=(len(cropped.rows),max((len(r) for r in cropped.rows),default=0)),issues=tuple(prefix+[merged_issue]),regions=regions,selected_region=chosen)
        return _inspect_flat_preview(cropped,source_path=source_path,source_format="xlsx",summaries=summaries,inspected_sheet=label,base_options=base,regions=regions,selected_region=chosen,prefix_issues=prefix)
    bounded=TablePreview(rows=tuple(tuple(row[:max_columns]) for row in preview.rows[:max_rows]),sheet_name=label,merged_ranges=preview.merged_ranges,apparent_rows=preview.apparent_rows,apparent_columns=preview.apparent_columns)
    return _inspect_flat_preview(bounded,source_path=source_path,source_format="xlsx",summaries=summaries,inspected_sheet=label,base_options=base,prefix_issues=prefix)
